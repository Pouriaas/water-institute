import numpy as np
import pandas as pd
import os
import sys
import openpyxl
from persiantools.jdatetime import JalaliDate
current_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(current_dir)
from shamsi_calender import shamsi_hourly

input_file = r'/home/pouria/git/water-institute/data/checking_similarity'



# Load calendar
shamsi_calen = pd.read_excel(r"/home/pouria/git/water-institute/data/solar_hijri_dates_1350_to_1401.xlsx")




files_and_directory = os.listdir(input_file)
for folder in files_and_directory:
    path_folder = os.path.join(input_file, folder)
    if not os.path.isdir(path_folder):
        continue  # skip files
    
    stations_name = os.listdir(path_folder)
    

    array = np.full((len(stations_name) + 1, len(shamsi_calen)), np.nan, dtype="object")
    df = pd.DataFrame(array.T, columns=["Date"] + stations_name)
    df["Date"] = shamsi_calen
    df = df.set_index("Date")
    # Output Excel file path
    output_excel_path = os.path.join(input_file,folder+".xlsx")
    # Open an ExcelWriter to write multiple sheets
    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:

        for file_name in stations_name:
            path = os.path.join(path_folder, file_name)
            if not file_name.endswith(".xlsx"):
                continue
            reame=file_name.replace("_hourly.xlsx","")
            reame=reame.replace("-","")
            wb = openpyxl.load_workbook(path)
            # in this time all the excel are green so we don't need remove by color
            # target_color = 'FF00B050'
            # purple_sheets = []
    
            # for sheetname in wb.sheetnames:
            #     sheet = wb[sheetname]
            #     tab_color = sheet.sheet_properties.tabColor
    
            #     if tab_color is not None:
            #         tab_color_rgb = tab_color.rgb
            #         if tab_color_rgb and tab_color_rgb.upper() == target_color:
            #             purple_sheets.append(sheetname)
    
            for sheet_name in wb.sheetnames:
                sim = pd.read_excel(path, sheet_name=sheet_name)
                valid_date = sim["Date"].dropna().unique()
                sim = sim.set_index("Date")
                for date in valid_date:
                    try:
                        if pd.isna(sim["Time"][date]):
                            continue
                        year = date.split("/")[0]
                        if year == "00":
                            year = "1400"
                        else:
                            year = "13" + str(year)
                        newformat = date.split("/")
                        new_date = f"{year}/{newformat[1]}/{newformat[2]}"
                        df.loc[new_date, file_name] = sheet_name
                    except:
                        continue  # Skip if indexing error
    
        
        # Define which rows qualify as part of an "event" (≥2 non-NaN values excluding 'Event')
        is_event_row = df.notna().sum(axis=1) >= 2
        
        # Mark consecutive blocks of event rows
        group_id = (is_event_row != is_event_row.shift()).cumsum()
        
        # Group only the rows that are part of event blocks
        grouped_events = df[is_event_row].groupby(group_id)
        
        # Build final DataFrame with labeled gaps
        for _, group in grouped_events:
            sim_list=[]
            start_date=100000000000
            end_date=0
            columnss=[]
            for column in group.columns:
                path_event=os.path.join(path_folder,column)
                valid_date = group[column].dropna().unique()
                if len(valid_date)!=0:
                    df_list = []
                    for event in valid_date:
                        sim = pd.read_excel(path_event, sheet_name=event)
                        # Fill NaN in existing 'Date'
                        sim['Date'] = sim['Date'].ffill()
                        df_list.append(sim)
                        
                    # Combine all dataframes in the list
                    final_df = pd.concat(df_list, ignore_index=True)
                    final_df=final_df.drop(columns=["STRTQ"])
                    first_last_date = final_df["Date"].dropna().unique()
                    
                    year = first_last_date[0].split("/")[0]
                    if year == "00":
                        year = "1400"
                    else:
                        year = "13" + str(year)
                    newformat = first_last_date[0].split("/")
                    newformat[0]=year
                    start=[int(i) for i in newformat]
                    start = JalaliDate(start[0], start[1], start[2]).toordinal()
                    
                    year = first_last_date[-1].split("/")[0]
                    if year == "00":
                        year = "1400"
                    else:
                        year = "13" + str(year)
                    newformat = first_last_date[-1].split("/")
                    newformat[0]=year
                    end=[int(i) for i in newformat]
                    end = JalaliDate(end[0], end[1], end[2]).toordinal()
                    
                    if start_date>start:
                        start_date=start
                    if end_date<end:
                        end_date=end
                    final_df['date'] = final_df['Date'].astype(str) + '/' + final_df['Time']
                    final_df=final_df.drop(columns=["Date","Time"]).set_index("date")
                    sim_list.append(final_df)
                    columnss.append(column)
            start_date=str(JalaliDate.fromordinal(start_date)).replace("-", "/")
            end_date=str(JalaliDate.fromordinal(end_date)).replace("-", "/")
            shamsi=shamsi_hourly(start_date,end_date)
            arr=np.full([len(shamsi),len(sim_list)+1],np.nan,dtype=object)
            df_all=pd.DataFrame(arr,columns=["Date"]+columnss)
            df_all.loc[:,"Date"]=shamsi
            df_all=df_all.set_index("Date")
            

            # Fill columns from df_list
            for i, df in enumerate(sim_list):
                # Align by index, extract the single column
                df_all[columnss[i]] = df.iloc[:, 0].reindex(df_all.index)
            df_all=df_all.reset_index()
            split_cols = df_all['Date'].str.rsplit('/', n=1, expand=True)
            split_cols.columns = ['date', 'time']  # Name the new columns
            
            # Reorder: put 'date' and 'time' at the front
            df_all = pd.concat([split_cols, df_all.drop(columns=['Date'])], axis=1)
            df_all = df_all.dropna(thresh=df_all.shape[1] - 1)
            # Sanitize the sheet name by replacing invalid characters
            sheet_name1 = str(df_all["date"].iloc[0])[:31].replace("/", "-")
            
            # Save the DataFrame to Excel with the sanitized sheet name
            df_all.to_excel(writer, sheet_name=sheet_name1,index=False)
            
            
            
                    

                        
                
                
            
            
            

        

        
