import os
import shutil
import subprocess
from openpyxl import load_workbook

import re
import pandas as pd
import openpyxl
import sys
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLabel, QLineEdit,
    QPushButton, QFileDialog, QMessageBox, QHBoxLayout
)

class HydrologicalApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Hydrological Model Automation")
        self.setGeometry(100, 100, 600, 250)

        layout = QVBoxLayout()

        # Source folder
        layout.addWidget(QLabel("Source Folder:"))
        self.source_input = QLineEdit()
        src_btn = QPushButton("Browse")
        src_btn.clicked.connect(self.browse_source)
        src_layout = QHBoxLayout()
        src_layout.addWidget(self.source_input)
        src_layout.addWidget(src_btn)
        layout.addLayout(src_layout)

        # Destination folder
        layout.addWidget(QLabel("Destination Folder:"))
        self.dest_input = QLineEdit()
        dst_btn = QPushButton("Browse")
        dst_btn.clicked.connect(self.browse_dest)
        dst_layout = QHBoxLayout()
        dst_layout.addWidget(self.dest_input)
        dst_layout.addWidget(dst_btn)
        layout.addLayout(dst_layout)

        # Sheets to delete
        layout.addWidget(QLabel("Sheet Names to Delete (/-separated):"))
        self.sheet_input = QLineEdit()
        layout.addWidget(self.sheet_input)

        # Run button
        run_btn = QPushButton("Run Automation")
        run_btn.clicked.connect(self.run_automation)
        layout.addWidget(run_btn)

        self.setLayout(layout)

        # Hardcoded model binary path (relative to this script)
        self.model_bin_path = os.path.join(os.path.dirname(__file__), 'model_bin')

    def browse_source(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Source Folder")
        if folder:
            self.source_input.setText(folder)

    def browse_dest(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Destination Folder")
        if folder:
            self.dest_input.setText(folder)

    def run_automation(self):
        src_folder = self.source_input.text().strip()
        dst_folder = self.dest_input.text().strip()
        
        sheets_to_delete = [s.strip() for s in self.sheet_input.text().split("/") if s.strip()]
        QMessageBox.information(self, "Success", f"{sheets_to_delete}")

        if not os.path.isdir(src_folder) or not os.path.isdir(dst_folder):
            QMessageBox.critical(self, "Error", "Please select valid folders.")
            return

        if not os.path.isdir(self.model_bin_path):
            QMessageBox.critical(self, "Error", f"Model binary folder not found:\n{self.model_bin_path}")
            return
        try:
            ##################################################################################
            # CONFIGURABLE
            source_folder = src_folder
            exefiles_folder = self.model_bin_path
            sheets_to_delete = sheets_to_delete  # <-- Replace with your actual sheet names
            dest_folder = dst_folder
            ##################################################################################



            os.makedirs(dest_folder, exist_ok=True)

            # Step 1: Copy folder + executables
            if os.path.exists(dest_folder):
                shutil.rmtree(dest_folder)
            shutil.copytree(source_folder, dest_folder)
            for fname in os.listdir(exefiles_folder):
                shutil.copy2(os.path.join(exefiles_folder, fname), dest_folder)

            # Find .xlsm after copying
            xlsm_list = [f for f in os.listdir(dest_folder) if f.endswith('.xlsm')]
            if not xlsm_list:
                raise FileNotFoundError("No .xlsm file found in source folder.")
            excel_path = os.path.join(dest_folder, xlsm_list[0])

            # Step 2: Load workbook & perform data processing
            wb = load_workbook(excel_path, data_only=True)
            sheets = wb.sheetnames
            for sheet in sheets_to_delete:
                if sheet in wb.sheetnames:
                    wb.remove(wb[sheet])

            sheets = wb.sheetnames
            QMessageBox.information(self, "Success", f"{sheets}")

            sheet_main = wb[sheets[0]]

            nEvent = len(sheets) - 1
            Tc = sheet_main.cell(row=3, column=1).value
            code = sheet_main.cell(row=1, column=1).value
            F_temp = sheet_main.cell(row=3, column=8).value

            Ar = [0] * 18
            Qq = [[0] * (nEvent + 1) for _ in range(300)]
            Rn = [[0] * (nEvent + 1) for _ in range(300)]
            Tp = [0] * 300
            STRTQ = [0] * (nEvent + 1)
            T_laps = [0] * (nEvent + 1)
            Dat = [""] * (nEvent + 1)
            Hou = [""] * (nEvent + 1)
            nn0 = [0] * (nEvent + 1)
            Area = 0

            ii = 2
            while True:
                ii += 1
                i = ii - 2
                val = sheet_main.cell(row=ii, column=5).value
                if not val or val == 0:
                    break
                Ar[i] = val
                Area += float(val)
            nAz = i - 1

            for j in range(1, nEvent + 1):
                sh = wb[sheets[j]]
                STRTQ[j] = sh.cell(row=3, column=5).value
                T_laps[j] = sh.cell(row=3, column=8).value
                Dat[j] = sh.cell(row=3, column=1).value
                Hou[j] = sh.cell(row=3, column=2).value

                ii, SumQ, SumR = 2, 0, 0
                while True:
                    ii += 1
                    i = ii - 2
                    Qv = sh.cell(row=ii, column=3).value
                    Rv = sh.cell(row=ii, column=4).value
                    if Qv is None and Rv is None:
                        break
                    qval = float(Qv or 0)
                    rval = float(Rv or 0)
                    Qq[i][j] = qval
                    Rn[i][j] = rval
                    SumQ += qval
                    SumR += rval

                n0 = i - 1
                nn0[j] = n0

                Sq0 = (SumQ - n0 * STRTQ[j]) * 3600
                sr0 = SumR * Area * 1000
                sheet_main.cell(row=j + 14, column=2).value = Sq0
                sheet_main.cell(row=j + 14, column=3).value = sr0
                sheet_main.cell(row=j + 14, column=4).value = Sq0 / sr0 if sr0 != 0 else None

                for i in range(1, n0 + 1):
                    sh.cell(row=i + 2, column=20).value = Qq[i][j]
                    sh.cell(row=i + 2, column=21).value = Rn[i][j]

                ii = 2
                while True:
                    ii += 1
                    i = ii - 2
                    Tpv = sh.cell(row=ii, column=7).value
                    if Tpv is None:
                        break
                    Tp[i] = float(Tpv)

                row_ptr = 2
                for k in range(1, n0 + 1):
                    aa = Tp[k]
                    for _ in range(24):
                        row_ptr += 1
                        sh.cell(row=row_ptr, column=9).value = aa
            xlsx_path=os.path.join(dest_folder, xlsm_list[0].replace('.xlsm', '_updated.xlsx'))
            wb.save(xlsx_path)

            # Step 3: Write .dat output files
            out_path = dest_folder + os.sep
            with open(out_path + "PSO_project.txt", "w") as f2:
                f2.write(f" ........  MaxIt nPop nVar  nEvent\n1600 600 6 {nEvent}\n")
                for tag in ["G", "R", "H", "T", "S"]:
                    f2.write(f".............................  {tag} Data file\n")
                    f2.write(out_path + f"Pso_{tag}_{code}.dat\n")

            def write_dat(filename, header_line, data_func):
                with open(out_path + filename, "w") as f:
                    for i in range(1, nEvent + 1):
                        f.write(f"{header_line} {i} :\n-\n")
                        f.write(f"   .............  Starting date and time for {filename[:-4]} data\n")
                        f.write(f"{Dat[i]}   {Hou[i]}\n")
                        f.write(f"   .............  n0 (number of records ) , Dt\n{nn0[i]} 1.0\n")
                        f.write(f"do i = 1 , n0          read  {filename[4].lower()} ( i )\n")
                        for j in range(1, nn0[i] + 1):
                            f.write(data_func(i, j) + "\n")

            write_dat(f"Pso_H_{code}.dat", "Event", lambda i, j: str(Qq[j][i]))
            write_dat(f"Pso_R_{code}.dat", "Event", lambda i, j: str(Rn[j][i]))
            write_dat(f"Pso_T_{code}.dat", "Event", lambda i, j: str(wb[sheets[i]].cell(row=j+2, column=9).value))

            # Write Pso_G_<code>.dat
            with open(out_path + f"Pso_G_{code}.dat", "w") as f:
                f.write("Area , Dt  nEvent\n")
                f.write(f"{Area:.3f} 1.0 {nEvent}\n")

                f.write(" Tc , Tc_min , Tc_max\n")
                f.write(f"{Tc:.3f} {0.7 * Tc:.3f} {1.3 * Tc:.3f}\n")

                f.write("Tr, Tr_min, Tr_max\n")
                f.write(" 1.0    .0       18.0\n")

                f.write("STRKR , _min , _max\n")
                f.write("3.09        .1       40.\n")

                f.write("DLTKR , _min , _max\n")
                f.write(" 4.172    0.1     8.0\n")

                f.write("RTIOL , _min , _max\n")
                f.write("2.277   1.0   3.0\n")

                f.write("ERAIN , _min , _max\n")
                f.write("0.727    .05   .95\n")

                f.write(" Freezing Temparature \n")
                f.write(f"{F_temp:.2f}\n")

                f.write("  STRTQ(i) .. T_laps(i) ...\n")
                for i in range(1, nEvent + 1):
                    f.write(f"{STRTQ[i]:.3f} {T_laps[i]:.3f}\n")

                f.write("No. of AreaZone\n")
                f.write(f"{nAz}\n")
                f.write(" Ar ( i ) , i = 1 , nAz \n")
                for i in range(nAz):
                    f.write(f"{Ar[i]:.3f}\n")

            # Step 4: Run the .exe
            exe_files = [f for f in os.listdir(dest_folder) if f.endswith('.exe')]
            if exe_files:
                exe_to_run = os.path.join(dest_folder, exe_files[0])
                result = subprocess.run(
                    [exe_to_run],
                    cwd=dest_folder,
                    creationflags=subprocess.CREATE_NEW_CONSOLE
                )
            else:
                print("No executable found to run.")
                
            folder_path=dest_folder
            input_excel_path = xlsx_path
            xls = pd.ExcelFile(input_excel_path)

            # Identify valid sheets
            sheet_names = [s for s in xls.sheet_names if s != "Geometry"]

            # Locate PSO file
            plot_candidates = [f for f in os.listdir(folder_path) if f.startswith("PSO_plot03")]
            if not plot_candidates:
                print("[!] No PSO_plot03 file found. Skipping.")
                sys.exit()
            plot_path = os.path.join(folder_path, plot_candidates[0])
            print(f"PSO file found: {plot_path}")

            # Read PSO file
            with open(plot_path, "r", encoding='utf-8', errors='ignore') as file:
                lines = file.readlines()

            # Prepare output writer
            output_path = os.path.join(folder_path, "output_plot03.xlsx")
            writer = pd.ExcelWriter(output_path, engine='openpyxl')
            data_frames = {}

            # Step 1: Build Observed column
            for sheet_name in sheet_names:
                df = xls.parse(sheet_name, header=None)
                df = df.iloc[1:].reset_index(drop=True)
                header_row = df.iloc[0]

                q_col_candidates = header_row[header_row.astype(str).str.strip().str.contains("Q", case=False, na=False)]
                if q_col_candidates.empty:
                    print(f"[!] Q column not found in sheet '{sheet_name}'. Skipping.")
                    continue

                q_col_index = q_col_candidates.index[0]
                observed_data = df[q_col_index].iloc[1:].reset_index(drop=True)
                output_df = pd.DataFrame({"Observed": observed_data})
                data_frames[sheet_name] = output_df

            # Step 2: Handle ZONE blocks
            current_zone = None
            current_data = []

            def process_zone(zone, data_lines):
                date_str = zone['date'].replace("/", ",")
                if date_str not in data_frames:
                    return

                df_sheet = xls.parse(date_str, header=None)
                df_sheet = df_sheet.iloc[1:].reset_index(drop=True)
                header_row = df_sheet.iloc[0]

                strtq_candidates = header_row[header_row.astype(str).str.strip().str.contains("STRTQ", case=False, na=False)]
                if strtq_candidates.empty:
                    print(f"[!] STRTQ not found in sheet '{date_str}'. Skipping zone.")
                    return

                strtq_index = strtq_candidates.index[0]
                strtq_value = df_sheet[strtq_index].iloc[1]

                new_column = []
                for row in data_lines:
                    parts = row.strip().split()
                    if len(parts) < 3:
                        break
                    try:
                        val2 = float(parts[1])
                        val3 = int(parts[2])
                        if val3 == 1:
                            new_column.append(val2 + float(strtq_value))
                    except:
                        continue

                col_name = f"S{zone['scenario']}E{zone['event']}"
                df_out = data_frames[date_str]
                df_out[col_name] = pd.Series(new_column)

            # Loop through PSO lines
            for line in lines:
                line = line.strip()
                if line.startswith("ZONE T="):
                    if current_zone:
                        process_zone(current_zone, current_data)
                    current_data = []
                    current_zone = {}
                    try:
                        date_part = line.split('"')[1].split()[0]
                        current_zone['date'] = date_part
                        scenario_match = re.search(r"Scenario=\s*(\d+)", line)
                        event_match = re.search(r"Event=\s*(\d+)", line)
                        if scenario_match and event_match:
                            current_zone['scenario'] = scenario_match.group(1)
                            current_zone['event'] = event_match.group(1)
                        else:
                            print(f"[!] Could not extract Scenario or Event in: {line}")
                            continue
                    except:
                        print(f"[!] Malformed ZONE line: {line}")
                        continue
                elif current_zone:
                    current_data.append(line)

            if current_zone:
                process_zone(current_zone, current_data)

            # Save output Excel
            for sheet_name, df in data_frames.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            writer.close()
            print(f"Done. Output saved to: {output_path}")

            #ocate PSO file
            plot_candidates = [f for f in os.listdir(folder_path) if f.startswith("PSO_plot04")]
            if not plot_candidates:
                print("[!] No PSO_plot04 file found. Skipping.")
                sys.exit()
            plot_path = os.path.join(folder_path, plot_candidates[0])
            print(f"PSO file found: {plot_path}")

            # Read PSO file
            with open(plot_path, "r", encoding='utf-8', errors='ignore') as file:
                lines = file.readlines()

            # Prepare output writer
            output_path = os.path.join(folder_path, "output_plot04.xlsx")
            writer = pd.ExcelWriter(output_path, engine='openpyxl')
            data_frames = {}

            # Step 1: Build Observed column
            for sheet_name in sheet_names:
                df = xls.parse(sheet_name, header=None)
                df = df.iloc[1:].reset_index(drop=True)
                header_row = df.iloc[0]

                q_col_candidates = header_row[header_row.astype(str).str.strip().str.contains("Q", case=False, na=False)]
                if q_col_candidates.empty:
                    print(f"[!] Q column not found in sheet '{sheet_name}'. Skipping.")
                    continue

                q_col_index = q_col_candidates.index[0]
                observed_data = df[q_col_index].iloc[1:].reset_index(drop=True)
                output_df = pd.DataFrame({"Observed": observed_data})
                data_frames[sheet_name] = output_df

            # Step 2: Handle ZONE blocks
            current_zone = None
            current_data = []

            def process_zone(zone, data_lines):
                date_str = zone['date'].replace("/", ",")
                if date_str not in data_frames:
                    return

                df_sheet = xls.parse(date_str, header=None)
                df_sheet = df_sheet.iloc[1:].reset_index(drop=True)
                header_row = df_sheet.iloc[0]

                strtq_candidates = header_row[header_row.astype(str).str.strip().str.contains("STRTQ", case=False, na=False)]
                if strtq_candidates.empty:
                    print(f"[!] STRTQ not found in sheet '{date_str}'. Skipping zone.")
                    return

                strtq_index = strtq_candidates.index[0]
                strtq_value = df_sheet[strtq_index].iloc[1]

                new_column = []
                for row in data_lines:
                    parts = row.strip().split()
                    if len(parts) < 3:
                        break
                    try:
                        val2 = float(parts[1])
                        val3 = int(parts[2])
                        if val3 == 1:
                            new_column.append(val2 + float(strtq_value))
                    except:
                        continue

                col_name = f"S{zone['scenario']}E{zone['event']}"
                df_out = data_frames[date_str]
                df_out[col_name] = pd.Series(new_column)

            # Loop through PSO lines
            for line in lines:
                line = line.strip()
                if line.startswith("ZONE T="):
                    if current_zone:
                        process_zone(current_zone, current_data)
                    current_data = []
                    current_zone = {}
                    try:
                        date_part = line.split('"')[1].split()[0]
                        current_zone['date'] = date_part
                        scenario_match = re.search(r"Scenario=\s*(\d+)", line)
                        event_match = re.search(r"Event=\s*(\d+)", line)
                        if scenario_match and event_match:
                            current_zone['scenario'] = scenario_match.group(1)
                            current_zone['event'] = event_match.group(1)
                        else:
                            print(f"[!] Could not extract Scenario or Event in: {line}")
                            continue
                    except:
                        print(f"[!] Malformed ZONE line: {line}")
                        continue
                elif current_zone:
                    current_data.append(line)

            if current_zone:
                process_zone(current_zone, current_data)

            # Save output Excel
            for sheet_name, df in data_frames.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            writer.close()
            print(f"Done. Output saved to: {output_path}")

            file1 = os.path.join(folder_path, "output_plot03.xlsx")
            file2 = os.path.join(folder_path, "output_plot04.xlsx")
            output_file = os.path.join(folder_path, "output.xlsx")

            if not os.path.exists(file1) or not os.path.exists(file2):
                print(f"Missing files in {folder_path}")
                sys.exit()


            excel1 = pd.read_excel(file1, sheet_name=None)
            excel2 = pd.read_excel(file2, sheet_name=None)

            writer = pd.ExcelWriter(output_file, engine='openpyxl')

            for sheet_name in excel1.keys():
                if sheet_name not in excel2:
                    continue

                df1 = excel1[sheet_name]
                df2 = excel2[sheet_name]

                if 'Observed' not in df1.columns or 'Observed' not in df2.columns:
                    print(f"Missing 'Observed' column in {sheet_name} of {folder_path}")
                    continue

                # Observed
                observed = df1['Observed'].reset_index(drop=True)

                # reversing
                df1_other = df1.drop(columns=['Observed']).iloc[::-1].reset_index(drop=True)
                df2_other = df2.drop(columns=['Observed']).iloc[::-1].reset_index(drop=True)

                
                final_data = pd.concat([observed, df1_other, df2_other], axis=1)

              
                top_row = [''] + ['New'] * df1_other.shape[1] + ['PSO'] * df2_other.shape[1]
                col_names = ['Observed'] + list(df1_other.columns) + list(df2_other.columns)

               
                final_data.columns = col_names

                
                top_df = pd.DataFrame([top_row], columns=col_names)
                final_data = pd.concat([top_df, final_data], ignore_index=True)

                
                final_data.to_excel(writer, sheet_name=sheet_name, index=False)

            writer.close()
            print(f"Processed: {folder_path}")


            QMessageBox.information(self, "Success", "All folders processed successfully.")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred:\n{str(e)}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = HydrologicalApp()
    window.show()
    sys.exit(app.exec())
