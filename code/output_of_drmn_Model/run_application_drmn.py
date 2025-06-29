# -*- coding: utf-8 -*-
"""
Created on Sun Jun 15 2025

@author: pouria
"""

import os
import shutil
import subprocess
from openpyxl import load_workbook

import re
import pandas as pd
import openpyxl
import sys

##################################################################################
# CONFIGURABLE
source_folder = r"C:/Users/pouria/Desktop/UTWI/DrMN/input/41"
exefiles_folder = r"C:/Users/pouria/Desktop/UTWI/DrMN/exefiles"
dest_main_folder = r"C:/Users/pouria/Desktop/UTWI/DrMN/output/41"
##################################################################################


def macro1(output_dir=None, output_excel=None, input_excel=None):


    # Load the workbook
    try:
        wb = openpyxl.load_workbook(input_excel)  # Replace with actual filename
    except FileNotFoundError:
        print("Error: Excel file not found. Please update the filename.")
        return
    """
    Python conversion of the VBA Macro1 subroutine for hydrological data processing
    """
    # Initialize arrays and variables
    Ar = [0] * 19  # VBA uses 1-based indexing, we'll use 0-based but keep one extra
    Qq = [[0] * 19 for _ in range(301)]  # 300x18 array
    Rn = [[0] * 19 for _ in range(301)]  # 300x18 array
    Tmp = [[0] * 19 for _ in range(301)]  # 300x18 array
    STRTQ = [0] * 19
    Dat = [""] * 19
    Hou = [""] * 19
    Tp = [0] * 301
    Tit = [""] * 19
    nn0 = [0] * 19
    T_laps = [0] * 19
    
    
    # Configuration parameters
    n_event = len(wb.sheetnames) - 1
    n_pop = 600
    max_it = 1600
    n_var = 6
    
    if n_event == 1:
        # Get the last sheet
        last_sheet = wb.sheetnames[-1]
        
        # Make a copy of the last sheet
        new_sheet = wb.create_sheet(title="99,99,99")
        
        # Copy all the data from the last sheet to the new sheet
        for row in wb[last_sheet].iter_rows():
            new_row = [cell.value for cell in row]
            new_sheet.append(new_row)
    
    # Read data from first sheet (index 0)
    n_event = len(wb.sheetnames) - 1
    sheet1 = wb.worksheets[0]
    
    Tc = sheet1.cell(3, 1).value or 0
    code = str(sheet1.cell(1, 1).value or "")
    F_temp = sheet1.cell(3, 8).value or 0
    
    # Read area data
    ii = 2
    Area = 0
    i = 0
    
    while True:
        ii += 1
        i = ii - 2
        if i >= len(Ar):
            break
        cell_value = sheet1.cell(ii, 5).value
        if cell_value is None or cell_value == 0:
            break
        Ar[i] = cell_value
        Area += Ar[i]
    
    nAz = i - 1
    
    # Process each event sheet
    for J in range(1, n_event + 1):
        jj = J  # Adjusted for 0-based indexing
        if jj >= len(wb.worksheets):
            break
            
        sheet = wb.worksheets[jj]
        
        # Read event parameters
        STRTQ[J] = sheet.cell(3, 5).value or 0
        T_laps[J] = sheet.cell(3, 8).value or 0
        Dat[J] = str(sheet.cell(3, 1).value or "")
        Hou[J] = str(sheet.cell(3, 2).value or "")
        
        # Read Q and R data
        ii = 2
        SumQ = 0
        SumR = 0
        
        while True:
            ii += 1
            i = ii - 2
            if i >= 300:  # Prevent array overflow
                break
                
            Q1_cell = sheet.cell(ii, 3).value
            r1_cell = sheet.cell(ii, 4).value
            
            if (Q1_cell is None or Q1_cell == "") and (r1_cell is None or r1_cell == ""):
                break
                
            Qq[i][J] = float(Q1_cell) if Q1_cell is not None else 0
            Rn[i][J] = float(r1_cell) if r1_cell is not None else 0
            SumQ += Qq[i][J]
            SumR += Rn[i][J]
        
        n0 = i - 1
        nn0[J] = n0
        
        # Calculate summary statistics
        Sq0 = (SumQ - n0 * STRTQ[J]) * 3600
        sr0 = SumR * Area * 1000
        
        # Write back to sheet1
        sheet1.cell(J + 14, 2).value = Sq0
        sheet1.cell(J + 14, 3).value = sr0
        if sr0 != 0:
            sheet1.cell(J + 14, 4).value = Sq0 / sr0
        
        # Copy data to columns 20-21
        for i in range(1, n0 + 1):
            ii = i + 2
            sheet.cell(ii, 20).value = Qq[i][J]
            sheet.cell(ii, 21).value = Rn[i][J]
        
        # Read temperature data
        ii = 2
        while True:
            ii += 1
            i = ii - 2
            if i >= 300:
                break
                
            Q1_cell = sheet.cell(ii, 7).value
            if Q1_cell is None or Q1_cell == "":
                break
                
            Tp[i] = float(Q1_cell) if Q1_cell is not None else 0
        
        ns = i - 1
        if ns > 0:
            a1 = Tp[1]
            
            # Interpolate temperature data
            ii = 2
            for i in range(1, n0 + 1):
                if i < len(Tp):
                    aa = Tp[i]
                    for k in range(1, 25):  # 1 to 24
                        ii += 1
                        if ii <= sheet.max_row:
                            sheet.cell(ii, 9).value = aa
    
    # Save the workbook
    wb.save(output_excel)
    

    
    # Create main project file
    create_pso_project_file(output_dir, max_it, n_pop, n_var, n_event, code)
    
    # Create rain-hydro data files
    create_rain_hydro_files(output_dir, n_event, nn0, Area, Rn, Qq, STRTQ)
    
    # Create additional data files
    create_general_data_file(output_dir, code, Area, n_event, Tc, F_temp, STRTQ, T_laps, nAz, Ar)
    create_hydrograph_file(output_dir, code, n_event, nn0, Dat, Hou, Qq)
    create_rain_file(output_dir, code, n_event, nn0, Dat, Hou, Rn)
    create_temperature_file(output_dir, code, n_event, nn0, Dat, Hou, wb)

def create_pso_project_file(output_dir, max_it, n_pop, n_var, n_event, code):
    """Create PSO_project.txt file"""
    with open(output_dir + "/" + "PSO_project.txt", "w") as f:
        f.write(" ........  MaxIt nPop nVar  nEvent\n")
        f.write(f"{max_it} {n_pop} {n_var} {n_event}\n")
        f.write(".............................  General Data file\n")
        f.write(f"Pso_G_{code}.dat\n")
        f.write(".............................  Rain Data file\n")
        f.write(f"Pso_R_{code}.dat\n")
        f.write(".............................  Hydrograph Data File\n")
        f.write(f"Pso_H_{code}.dat\n")
        f.write(".............................  Temparature Data File\n")
        f.write(f"Pso_T_{code}.dat\n")
        f.write(".............................  Snow Data File\n")
        f.write(f"Pso_S_{code}.dat\n")

def create_rain_hydro_files(output_dir, n_event, nn0, Area, Rn, Qq, STRTQ):
    """Create Rain_Hydro files"""
    for J in range(1, n_event + 1):
        filename = f"Rain_Hydro_16211_{chr(J + 64)}.Dat"
        with open(output_dir + "/" + filename, "w") as f:
            n0 = nn0[J]
            for i in range(1, n0 + 1):
                rain_val = Area * Rn[i][J] / 3.6
                flow_val = Qq[i][J] - STRTQ[J]
                f.write(f"{i} {rain_val} {flow_val}\n")

def create_general_data_file(output_dir, code, Area, n_event, Tc, F_temp, STRTQ, T_laps, nAz, Ar):
    """Create general data file"""
    with open(output_dir + "/" + f"Pso_G_{code}.dat", "w") as f:
        f.write("Area , Dt  nEvent\n")
        f.write(f"{Area} 1.0 {n_event}\n")
        f.write(" Tc , Tc_min , Tc_max\n")
        f.write(f"{Tc} {0.7 * Tc} {1.3 * Tc}\n")
        f.write("Tr, Tr_min, Tr_max\n")
        f.write(" 1.0    .0       18.0\n")
        f.write("STRKR , _min , _max\n")
        f.write("3.09        .1       40.\n")
        f.write("DLTKR , _min , _max\n")
        f.write(" 4.172    0.1     8.0\n")
        f.write("RTIOL , _min , _max\n")
        f.write("2.277   1.0   3.0\n")
        f.write("ERAIN , _min , _max\n")
        f.write("0.727    .05   .95\n")
        f.write(" Freezing Temparature \n")
        f.write(f"{F_temp}\n")
        f.write("  STRTQ(i) .. T_laps(i) ...\n")
        
        for l0 in range(1, n_event + 1):
            f.write(f"{STRTQ[l0]} {T_laps[l0]}\n")
        
        f.write("No. of AreaZone\n")
        f.write(f"{nAz}\n")
        f.write(" Ar ( i ) , i = 1 , nAz \n")
        
        for l0 in range(1, nAz + 1):
            f.write(f"{Ar[l0]}\n")

def create_hydrograph_file(output_dir, code, n_event, nn0, Dat, Hou, Qq):
    """Create hydrograph data file"""
    titles = ["Fisrt Event :", "Second Event :", "Third Event :", "4th Event :", 
              "5th Event :", "6th Event :", "7th Event :", "8th Event :", 
              "9th Event :", "10th Event :", "11th Event :"]
    
    with open(output_dir + "/" + f"Pso_H_{code}.dat", "w") as f:
        for i in range(1, n_event + 1):
            n0 = nn0[i]
            title = titles[i-1] if i-1 < len(titles) else f"{i}th Event :"
            
            f.write(f"{title}\n")
            f.write("-\n")
            f.write("   .............  Starting date and time for Hydrograph data\n")
            f.write(f"{Dat[i]}   {Hou[i]}\n")
            f.write("   .............  n0 (number of records ) , Dt\n")
            f.write(f"{n0} 1.0\n")
            f.write("do i = 1 , n0          read  Qq ( i )\n")
            
            for J in range(1, n0 + 1):
                f.write(f"{Qq[J][i]}\n")

def create_rain_file(output_dir, code, n_event, nn0, Dat, Hou, Rn):
    """Create rain data file"""
    titles = ["Fisrt Event :", "Second Event :", "Third Event :", "Fourth Event :", "Fifth Event :"]
    
    with open(output_dir + "/" + f"Pso_R_{code}.dat", "w") as f:
        for i in range(1, n_event + 1):
            n0 = nn0[i]
            title = titles[i-1] if i-1 < len(titles) else f"{i}th Event :"
            
            f.write(f"{title}\n")
            f.write("-\n")
            f.write("   .............  Starting date and time for rain data\n")
            f.write(f"{Dat[i]}   {Hou[i]}\n")
            f.write("   .............  n0 (number of records ) , Dt\n")
            f.write(f"{n0} 1.0\n")
            f.write("do i = 1 , n0          read  rn ( i )\n")
            
            for J in range(1, n0 + 1):
                f.write(f"{Rn[J][i]}\n")

def create_temperature_file(output_dir, code, n_event, nn0, Dat, Hou, wb):
    """Create temperature data file"""
    titles = ["Fisrt Event :", "Second Event :", "Third Event :", "Fourth Event :", "Fifth Event :"]
    
    with open(output_dir + "/" + f"Pso_T_{code}.dat", "w") as f:
        for i in range(1, n_event + 1):
            n0 = nn0[i]
            title = titles[i-1] if i-1 < len(titles) else f"{i}th Event :"
            
            f.write(f"{title}\n")
            f.write("-\n")
            f.write("   .............  Starting date and time for Temparature data\n")
            f.write(f"{Dat[i]}   {Hou[i]}\n")
            f.write("   .............  n0 (number of records ) , Dt\n")
            f.write(f"{n0} 1.0\n")
            f.write("do i = 1 , n0          read  rn ( i )\n")
            
            # Read temperature data from the corresponding sheet
            if i < len(wb.worksheets):
                sheet = wb.worksheets[i]
                for J in range(1, n0 + 1):
                    ii = J + 2
                    tmp0 = sheet.cell(ii, 9).value or 0
                    f.write(f"{tmp0}\n")

stations=os.listdir(source_folder)

for files in stations:
    
    dest_folder=os.path.join(dest_main_folder, files[:6])
    os.makedirs(dest_folder, exist_ok=True)
    shutil.copy2(os.path.join(source_folder, files), dest_folder)
    # Step 1: Copy folder + executables

    for fname in os.listdir(exefiles_folder):
        shutil.copy2(os.path.join(exefiles_folder, fname), dest_folder)
    # Find .xlsm after copying

    excel_path = os.path.join(dest_folder, files)
    
    # Step 2: Load workbook & perform data processing
    macro1(output_dir=dest_folder, output_excel=excel_path, input_excel=excel_path)
    
    # Step 4: Run the .exe
    exe_files = [f for f in os.listdir(dest_folder) if f.endswith('.exe')]
    if exe_files:
        exe_to_run = os.path.join(dest_folder, exe_files[0])
        result = subprocess.run(
            [exe_to_run],
            cwd=dest_folder,
            creationflags=subprocess.CREATE_NEW_CONSOLE
        )
    else:
        print("No executable found to run.")
    
    folder_path=dest_folder
    input_excel_path = excel_path
    xls = pd.ExcelFile(input_excel_path)
    
    # Identify valid sheets
    sheet_names = [s for s in xls.sheet_names if s != "Geometry"]
    
    # Locate PSO file
    plot_candidates = [f for f in os.listdir(folder_path) if f.startswith("PSO_plot03")]
    if not plot_candidates:
        print("[!] No PSO_plot03 file found. Skipping.")
        sys.exit()
    plot_path = os.path.join(folder_path, plot_candidates[0])
    print(f"PSO file found: {plot_path}")
    
    # Read PSO file
    with open(plot_path, "r", encoding='utf-8', errors='ignore') as file:
        lines = file.readlines()
    
    # Prepare output writer
    output_path = os.path.join(folder_path, "output_plot03.xlsx")
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    data_frames = {}
    
    # Step 1: Build Observed column
    for sheet_name in sheet_names:
        df = xls.parse(sheet_name, header=None)
        df = df.iloc[1:].reset_index(drop=True)
        header_row = df.iloc[0]
    
        q_col_candidates = header_row[header_row.astype(str).str.strip().str.contains("Q", case=False, na=False)]
        if q_col_candidates.empty:
            print(f"[!] Q column not found in sheet '{sheet_name}'. Skipping.")
            continue
    
        q_col_index = q_col_candidates.index[0]
        observed_data = df[q_col_index].iloc[1:].reset_index(drop=True)
        output_df = pd.DataFrame({"Observed": observed_data})
        data_frames[sheet_name] = output_df
    
    # Step 2: Handle ZONE blocks
    current_zone = None
    current_data = []
    
    def process_zone(zone, data_lines):
        date_str = zone['date'].replace("/", ",")
        if date_str not in data_frames:
            return
    
        df_sheet = xls.parse(date_str, header=None)
        df_sheet = df_sheet.iloc[1:].reset_index(drop=True)
        header_row = df_sheet.iloc[0]
    
        strtq_candidates = header_row[header_row.astype(str).str.strip().str.contains("STRTQ", case=False, na=False)]
        if strtq_candidates.empty:
            print(f"[!] STRTQ not found in sheet '{date_str}'. Skipping zone.")
            return
    
        strtq_index = strtq_candidates.index[0]
        strtq_value = df_sheet[strtq_index].iloc[1]
    
        new_column = []
        for row in data_lines:
            parts = row.strip().split()
            if len(parts) < 3:
                break
            try:
                val2 = float(parts[1])
                val3 = int(parts[2])
                if val3 == 1:
                    new_column.append(val2 + float(strtq_value))
            except:
                continue
    
        col_name = f"S{zone['scenario']}E{zone['event']}"
        df_out = data_frames[date_str]
        df_out[col_name] = pd.Series(new_column)
    
    # Loop through PSO lines
    for line in lines:
        line = line.strip()
        if line.startswith("ZONE T="):
            if current_zone:
                process_zone(current_zone, current_data)
            current_data = []
            current_zone = {}
            try:
                date_part = line.split('"')[1].split()[0]
                current_zone['date'] = date_part
                scenario_match = re.search(r"Scenario=\s*(\d+)", line)
                event_match = re.search(r"Event=\s*(\d+)", line)
                if scenario_match and event_match:
                    current_zone['scenario'] = scenario_match.group(1)
                    current_zone['event'] = event_match.group(1)
                else:
                    print(f"[!] Could not extract Scenario or Event in: {line}")
                    continue
            except:
                print(f"[!] Malformed ZONE line: {line}")
                continue
        elif current_zone:
            current_data.append(line)
    
    if current_zone:
        process_zone(current_zone, current_data)
    
    # Save output Excel
    for sheet_name, df in data_frames.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    writer.close()
    print(f"Done. Output saved to: {output_path}")
    
    #ocate PSO file
    plot_candidates = [f for f in os.listdir(folder_path) if f.startswith("PSO_plot04")]
    if not plot_candidates:
        print("[!] No PSO_plot04 file found. Skipping.")
        sys.exit()
    plot_path = os.path.join(folder_path, plot_candidates[0])
    print(f"PSO file found: {plot_path}")
    
    # Read PSO file
    with open(plot_path, "r", encoding='utf-8', errors='ignore') as file:
        lines = file.readlines()
    
    # Prepare output writer
    output_path = os.path.join(folder_path, "output_plot04.xlsx")
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    data_frames = {}
    
    # Step 1: Build Observed column
    for sheet_name in sheet_names:
        df = xls.parse(sheet_name, header=None)
        df = df.iloc[1:].reset_index(drop=True)
        header_row = df.iloc[0]
    
        q_col_candidates = header_row[header_row.astype(str).str.strip().str.contains("Q", case=False, na=False)]
        if q_col_candidates.empty:
            print(f"[!] Q column not found in sheet '{sheet_name}'. Skipping.")
            continue
    
        q_col_index = q_col_candidates.index[0]
        observed_data = df[q_col_index].iloc[1:].reset_index(drop=True)
        output_df = pd.DataFrame({"Observed": observed_data})
        data_frames[sheet_name] = output_df
    
    # Step 2: Handle ZONE blocks
    current_zone = None
    current_data = []
    
    def process_zone(zone, data_lines):
        date_str = zone['date'].replace("/", ",")
        if date_str not in data_frames:
            return
    
        df_sheet = xls.parse(date_str, header=None)
        df_sheet = df_sheet.iloc[1:].reset_index(drop=True)
        header_row = df_sheet.iloc[0]
    
        strtq_candidates = header_row[header_row.astype(str).str.strip().str.contains("STRTQ", case=False, na=False)]
        if strtq_candidates.empty:
            print(f"[!] STRTQ not found in sheet '{date_str}'. Skipping zone.")
            return
    
        strtq_index = strtq_candidates.index[0]
        strtq_value = df_sheet[strtq_index].iloc[1]
    
        new_column = []
        for row in data_lines:
            parts = row.strip().split()
            if len(parts) < 3:
                break
            try:
                val2 = float(parts[1])
                val3 = int(parts[2])
                if val3 == 1:
                    new_column.append(val2 + float(strtq_value))
            except:
                continue
    
        col_name = f"S{zone['scenario']}E{zone['event']}"
        df_out = data_frames[date_str]
        df_out[col_name] = pd.Series(new_column)
    
    # Loop through PSO lines
    for line in lines:
        line = line.strip()
        if line.startswith("ZONE T="):
            if current_zone:
                process_zone(current_zone, current_data)
            current_data = []
            current_zone = {}
            try:
                date_part = line.split('"')[1].split()[0]
                current_zone['date'] = date_part
                scenario_match = re.search(r"Scenario=\s*(\d+)", line)
                event_match = re.search(r"Event=\s*(\d+)", line)
                if scenario_match and event_match:
                    current_zone['scenario'] = scenario_match.group(1)
                    current_zone['event'] = event_match.group(1)
                else:
                    print(f"[!] Could not extract Scenario or Event in: {line}")
                    continue
            except:
                print(f"[!] Malformed ZONE line: {line}")
                continue
        elif current_zone:
            current_data.append(line)
    
    if current_zone:
        process_zone(current_zone, current_data)
    
    # Save output Excel
    for sheet_name, df in data_frames.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    writer.close()
    print(f"Done. Output saved to: {output_path}")
    
    file1 = os.path.join(folder_path, "output_plot03.xlsx")
    file2 = os.path.join(folder_path, "output_plot04.xlsx")
    output_file = os.path.join(folder_path, "output.xlsx")
    
    if not os.path.exists(file1) or not os.path.exists(file2):
        print(f"Missing files in {folder_path}")
        sys.exit()
    
    
    excel1 = pd.read_excel(file1, sheet_name=None)
    excel2 = pd.read_excel(file2, sheet_name=None)
    
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    
    for sheet_name in excel1.keys():
        if sheet_name not in excel2:
            continue
    
        df1 = excel1[sheet_name]
        df2 = excel2[sheet_name]
    
        if 'Observed' not in df1.columns or 'Observed' not in df2.columns:
            print(f"Missing 'Observed' column in {sheet_name} of {folder_path}")
            continue
    
        # Observed
        observed = df1['Observed'].reset_index(drop=True)
    
        # reversing
        df1_other = df1.drop(columns=['Observed']).iloc[::-1].reset_index(drop=True)
        df2_other = df2.drop(columns=['Observed']).iloc[::-1].reset_index(drop=True)
    
        
        final_data = pd.concat([observed, df1_other, df2_other], axis=1) 
    
      
        top_row = [''] + ['New'] * df1_other.shape[1] + ['PSO'] * df2_other.shape[1]
        col_names = ['Observed'] + list(df1_other.columns) + list(df2_other.columns)
    
       
        final_data.columns = col_names
    
        
        top_df = pd.DataFrame([top_row], columns=col_names)
        final_data = pd.concat([top_df, final_data], ignore_index=True)
    
        
        final_data.to_excel(writer, sheet_name=sheet_name, index=False)
    
    writer.close()
    print(f"Processed: {folder_path}")
