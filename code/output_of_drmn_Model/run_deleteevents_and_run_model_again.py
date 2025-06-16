# -*- coding: utf-8 -*-
"""
Created on Sun Jun 15 2025

@author: pouria
"""

import os
import shutil
import subprocess
from openpyxl import load_workbook

##################################################################################
# CONFIGURABLE
source_folder = r"C:/Users/pouria/Desktop/UTWI/DrMN/16-019"
exefiles_folder = r"C:/Users/pouria/Desktop/UTWI/DrMN/exefiles"
folder_name = "941210"
sheets_to_delete = ['94,12,10']  # <-- Replace with your actual sheet names
dest_folder = os.path.join(r"C:/Users/pouria/Desktop/UTWI/DrMN", folder_name)
##################################################################################



os.makedirs(dest_folder, exist_ok=True)

# Step 1: Copy folder + executables
if os.path.exists(dest_folder):
    shutil.rmtree(dest_folder)
shutil.copytree(source_folder, dest_folder)
for fname in os.listdir(exefiles_folder):
    shutil.copy2(os.path.join(exefiles_folder, fname), dest_folder)

# Find .xlsm after copying
xlsm_list = [f for f in os.listdir(dest_folder) if f.endswith('.xlsm')]
if not xlsm_list:
    raise FileNotFoundError("No .xlsm file found in source folder.")
excel_path = os.path.join(dest_folder, xlsm_list[0])

# Step 2: Load workbook & perform data processing
wb = load_workbook(excel_path, data_only=True)
sheets = wb.sheetnames
for sheet in sheets_to_delete:
    if sheet in wb.sheetnames:
        wb.remove(wb[sheet])

sheets = wb.sheetnames
sheet_main = wb[sheets[0]]

nEvent = len(sheets) - 1
Tc = sheet_main.cell(row=3, column=1).value
code = sheet_main.cell(row=1, column=1).value
F_temp = sheet_main.cell(row=3, column=8).value

Ar = [0] * 18
Qq = [[0] * (nEvent + 1) for _ in range(300)]
Rn = [[0] * (nEvent + 1) for _ in range(300)]
Tp = [0] * 300
STRTQ = [0] * (nEvent + 1)
T_laps = [0] * (nEvent + 1)
Dat = [""] * (nEvent + 1)
Hou = [""] * (nEvent + 1)
nn0 = [0] * (nEvent + 1)
Area = 0

ii = 2
while True:
    ii += 1
    i = ii - 2
    val = sheet_main.cell(row=ii, column=5).value
    if not val or val == 0:
        break
    Ar[i] = val
    Area += float(val)
nAz = i - 1

for j in range(1, nEvent + 1):
    sh = wb[sheets[j]]
    STRTQ[j] = sh.cell(row=3, column=5).value
    T_laps[j] = sh.cell(row=3, column=8).value
    Dat[j] = sh.cell(row=3, column=1).value
    Hou[j] = sh.cell(row=3, column=2).value

    ii, SumQ, SumR = 2, 0, 0
    while True:
        ii += 1
        i = ii - 2
        Qv = sh.cell(row=ii, column=3).value
        Rv = sh.cell(row=ii, column=4).value
        if Qv is None and Rv is None:
            break
        qval = float(Qv or 0)
        rval = float(Rv or 0)
        Qq[i][j] = qval
        Rn[i][j] = rval
        SumQ += qval
        SumR += rval

    n0 = i - 1
    nn0[j] = n0

    Sq0 = (SumQ - n0 * STRTQ[j]) * 3600
    sr0 = SumR * Area * 1000
    sheet_main.cell(row=j + 14, column=2).value = Sq0
    sheet_main.cell(row=j + 14, column=3).value = sr0
    sheet_main.cell(row=j + 14, column=4).value = Sq0 / sr0 if sr0 != 0 else None

    for i in range(1, n0 + 1):
        sh.cell(row=i + 2, column=20).value = Qq[i][j]
        sh.cell(row=i + 2, column=21).value = Rn[i][j]

    ii = 2
    while True:
        ii += 1
        i = ii - 2
        Tpv = sh.cell(row=ii, column=7).value
        if Tpv is None:
            break
        Tp[i] = float(Tpv)

    row_ptr = 2
    for k in range(1, n0 + 1):
        aa = Tp[k]
        for _ in range(24):
            row_ptr += 1
            sh.cell(row=row_ptr, column=9).value = aa

wb.save(os.path.join(dest_folder, xlsm_list[0].replace('.xlsm', '_updated.xlsx')))

# Step 3: Write .dat output files
out_path = dest_folder + os.sep
with open(out_path + "PSO_project.txt", "w") as f2:
    f2.write(f" ........  MaxIt nPop nVar  nEvent\n1600 600 6 {nEvent}\n")
    for tag in ["G", "R", "H", "T", "S"]:
        f2.write(f".............................  {tag} Data file\n")
        f2.write(out_path + f"Pso_{tag}_{code}.dat\n")

def write_dat(filename, header_line, data_func):
    with open(out_path + filename, "w") as f:
        for i in range(1, nEvent + 1):
            f.write(f"{header_line} {i} :\n-\n")
            f.write(f"   .............  Starting date and time for {filename[:-4]} data\n")
            f.write(f"{Dat[i]}   {Hou[i]}\n")
            f.write(f"   .............  n0 (number of records ) , Dt\n{nn0[i]} 1.0\n")
            f.write(f"do i = 1 , n0          read  {filename[4].lower()} ( i )\n")
            for j in range(1, nn0[i] + 1):
                f.write(data_func(i, j) + "\n")

write_dat(f"Pso_H_{code}.dat", "Event", lambda i, j: str(Qq[j][i]))
write_dat(f"Pso_R_{code}.dat", "Event", lambda i, j: str(Rn[j][i]))
write_dat(f"Pso_T_{code}.dat", "Event", lambda i, j: str(wb[sheets[i]].cell(row=j+2, column=9).value))

# Write Pso_G_<code>.dat
with open(out_path + f"Pso_G_{code}.dat", "w") as f:
    f.write("Area , Dt  nEvent\n")
    f.write(f"{Area:.3f} 1.0 {nEvent}\n")

    f.write(" Tc , Tc_min , Tc_max\n")
    f.write(f"{Tc:.3f} {0.7 * Tc:.3f} {1.3 * Tc:.3f}\n")

    f.write("Tr, Tr_min, Tr_max\n")
    f.write(" 1.0    .0       18.0\n")

    f.write("STRKR , _min , _max\n")
    f.write("3.09        .1       40.\n")

    f.write("DLTKR , _min , _max\n")
    f.write(" 4.172    0.1     8.0\n")

    f.write("RTIOL , _min , _max\n")
    f.write("2.277   1.0   3.0\n")

    f.write("ERAIN , _min , _max\n")
    f.write("0.727    .05   .95\n")

    f.write(" Freezing Temparature \n")
    f.write(f"{F_temp:.2f}\n")

    f.write("  STRTQ(i) .. T_laps(i) ...\n")
    for i in range(1, nEvent + 1):
        f.write(f"{STRTQ[i]:.3f} {T_laps[i]:.3f}\n")

    f.write("No. of AreaZone\n")
    f.write(f"{nAz}\n")
    f.write(" Ar ( i ) , i = 1 , nAz \n")
    for i in range(nAz):
        f.write(f"{Ar[i]:.3f}\n")

# Step 4: Run the .exe
exe_files = [f for f in os.listdir(dest_folder) if f.endswith('.exe')]
if exe_files:
    exe_to_run = os.path.join(dest_folder, exe_files[0])
    subprocess.Popen(
        [exe_to_run],
        cwd=dest_folder,
        creationflags=subprocess.CREATE_NEW_CONSOLE
    )
else:
    print("No executable found to run.")