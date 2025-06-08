#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Jun  8 07:48:30 2025

@author: pouria
"""

# Import required libraries
import numpy as np
import pandas as pd
import os
import math
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Font

# Define the folder containing the Excel files to modify
input_folder = r"/home/pouria/git/water-institute/data/drmn_input/input"

# Load reference Excel files used for enriching data
point_tp = pd.read_excel(r"/home/pouria/git/water-institute/data/drmn_input/12 PointTemperature/12 PointTemperature-4040128.xlsx", sheet_name="Sheet1").set_index("Point_ID")
hydro_attribute = pd.read_excel(r"/home/pouria/git/water-institute/data/drmn_input/Hydro_Attribute Table.xlsx").set_index("Code")
characteristic = pd.read_excel(r"/home/pouria/git/water-institute/data/drmn_input/12_SubWatershed_Characteristics_HydroStations/12_Talesh_Anzali_SubWatershed_Characteristics_HydroStations_4040125.xlsx", sheet_name="Sheet1", skiprows=2).set_index("Point_ID")
snow = pd.read_excel(r"I:\WMS_Project\14 Lahijan-Noor\14 Characteristics\14 Snow Parameters.xlsx", sheet_name="Sheet1").set_index("Basin_3rd")

# Define folders for supplementary input data
hypsometric_folder = r"/home/pouria/git/water-institute/data/drmn_input/Hypsometric_Tables"
TemperatureGradientfolder = r"/home/pouria/git/water-institute/data/drmn_input/12 TemperatureGradient"
files_and_directory = os.listdir(input_folder)

# Process each Excel file in the input folder
for name in files_and_directory:
    path = os.path.join(input_folder, name)
    wb = load_workbook(path)  # Load workbook while preserving charts
    rename = name[:2] + name[3:6]  # Generate identifier from file name

    # Load hypsometric data for the subwatershed
    hypsometric_path = os.path.join(hypsometric_folder, rename + '.xlsx')
    hypsometric = pd.read_excel(hypsometric_path)

    # Extract various characteristics for the watershed into a dictionary
    sim_data = {
        "Basin_3rd": hydro_attribute.loc[int(rename), "OID"],
        "POINT_X": hydro_attribute.loc[int(rename), "XG"],
        "POINT_Y": hydro_attribute.loc[int(rename), "YG"],
        "Tc (hr)": characteristic.loc[int(rename), "Tc(Krp_hr)"],
        "Total Area (km2)": hypsometric.loc[0, "Cumulative Area (km²)"],
        "Snow melt coeff": snow.loc[hydro_attribute.loc[int(rename), "Basin3_ID"], "Snow melt coeff"],
        "Freezing temp (C)": snow.loc[hydro_attribute.loc[int(rename), "Basin3_ID"], "Freezing temp (C)"]
    }

    # Load temperature gradient table for the basin
    TemperatureGradientpath = os.path.join(TemperatureGradientfolder, str(sim_data["Basin_3rd"]) + ".xlsx")
    TemperatureGradient = pd.read_excel(TemperatureGradientpath)

    # Preprocess elevation ranges to group them
    hypsometric[['Start', 'End']] = hypsometric['Elevation Range (m)'].str.split(' - ', expand=True).astype(float)
    hypsometric = hypsometric.sort_values('Start').reset_index(drop=True)
    num_rows = len(hypsometric)
    num_groups = min(10, num_rows)
    group_size = math.ceil(num_rows / num_groups)
    hypsometric['Group'] = hypsometric.index // group_size

    # Aggregate elevation data by group
    grouped = hypsometric.groupby('Group').agg({
        'Start': 'min',
        'End': 'max',
        'Area (km²)': 'sum'
    }).reset_index(drop=True)

    # Add zone height data to sim_data
    sim_data["Hydrometry Station Elevation"] = grouped.loc[0, 'Start']
    sim_data["Zone height (m)"] = grouped.loc[0, 'End'] - grouped.loc[0, 'Start']

    # Modify each sheet in the workbook
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Convert worksheet to DataFrame
        df = pd.DataFrame(ws.values)
        df.columns = df.iloc[0]  # Set header
        df = df.drop(index=0).reset_index(drop=True)
        df['Date'] = df['Date'].ffill()  # Fill forward missing dates

        # Extract unique filled dates
        new_dates = df['Date'].dropna().unique()
        temperature_values = []

        # Compute temperature from point_tp using the date
        for datee in new_dates:
            year = datee.split("/")[0]
            datee = datee.replace("/", "")
            datee = str("14" + datee) if int(year) < 3 else str("13" + datee)
            temperature_values.append(point_tp.loc[int(rename), int(datee)])

        # Write new date and temperature to new columns (append to right)
        max_col=ws.max_column
        # Define border style (thin black lines)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Define bold font
        bold_font = Font(bold=True)
        
        # Apply to headers
        for col_offset, header in enumerate(["Date", "Temperature", "temp.lapse/zone"], start=1):
            cell = ws.cell(row=1, column=max_col + col_offset, value=header)
            cell.border = thin_border
            cell.font = bold_font
        for idx, val in enumerate(new_dates):
            ws.cell(row=idx+2, column=max_col + 1, value=val)  # Write new date
            ws.cell(row=idx+2, column=max_col + 2, value=temperature_values[idx])  # Overwrites same column!

        # Compute temperature lapse rate from gradient and elevation
        month = int(str(df.loc[0, 'Date']).split("/")[1])
        formula = TemperatureGradient.iloc[month - 1, 1]
        match = abs(float(re.search(r'([-+]?\d*\.?\d+)\*H', formula).group(1)))
        ws.cell(row=2, column=ws.max_column, value=match * sim_data["Zone height (m)"])

    # Save workbook to the same file, preserving plots


    # Define correct column order
    ordered_columns = ["Tc (hr)", "Hydrometry Station Elevation", "Elevation Range (m) strat",
                       "Elevation Range (m) end", "Area (km2)", "Zone height (m)",
                       "Snow melt coeff", "Freezing temp (C)", "Basin_3rd", 
                       "Total Area (km2)", "POINT_X", "POINT_Y"]
    
    # Initialize DataFrame for sim (geometry table) with proper column order
    sim = pd.DataFrame(np.full((10, len(ordered_columns)), np.nan), columns=ordered_columns)
    
    # Fill values using your existing logic
    sim.loc[0, "Basin_3rd"] = sim_data["Basin_3rd"]
    sim.loc[0, "POINT_X"] = sim_data["POINT_X"]
    sim.loc[0, "POINT_Y"] = sim_data["POINT_Y"]
    sim.loc[0, "Tc (hr)"] = sim_data["Tc (hr)"]
    sim.loc[0, "Total Area (km2)"] = sim_data["Total Area (km2)"]
    sim.loc[0, "Snow melt coeff"] = sim_data["Snow melt coeff"]
    sim.loc[0, "Freezing temp (C)"] = sim_data["Freezing temp (C)"]
    sim.loc[0, "Hydrometry Station Elevation"] = sim_data["Hydrometry Station Elevation"]
    sim.loc[0, "Zone height (m)"] = sim_data["Zone height (m)"]
    
    # Add grouped elevation data
    sim.loc[:len(grouped)-1, "Elevation Range (m) strat"] = grouped["Start"].values
    sim.loc[:len(grouped)-1, "Elevation Range (m) end"] = grouped["End"].values
    sim.loc[:len(grouped)-1, "Area (km2)"] = grouped["Area (km²)"].values
    
    # Add "Geometry" worksheet
    if "Geometry" in wb.sheetnames:
        del wb["Geometry"]  # Remove existing to overwrite
    ws_geometry = wb.create_sheet("Geometry", 0)  # Insert at first position
    
    # Write sim DataFrame to new sheet
    for row in dataframe_to_rows(sim[ordered_columns], index=False, header=True):
        ws_geometry.append(row)
    for cell in ws_geometry[1]:  # Row 1 is the header
        cell.font = bold_font
        cell.border = thin_border
    wb.save(path)

print("Done!")