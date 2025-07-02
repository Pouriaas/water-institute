#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Jul  2 14:08:00 2025

@author: pouria
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate and run one simulation per Excel file using 'True' parameter values — parallelized.
"""

import os
import shutil
import subprocess
import pandas as pd
from openpyxl import load_workbook
from joblib import Parallel, delayed
import multiprocessing
from tqdm import tqdm

# === CONFIGURATION ===
excel_dir = "C:/Users/DrBZahraei/Desktop/regenalization/hydrographs"
exefiles_dir = "C:/Users/DrBZahraei/Desktop/regenalization/Edit-Hec-For_Dr_Zahraii2"
output_base_dir = "C:/Users/DrBZahraei/Desktop/regenalization/output"
param_master_file = "C:/Users/DrBZahraei/Desktop/regenalization/parameters.xlsx"
exe_name = "Read_input_xlsx_+_observ.exe"
parameters = ['Tr', 'STRKR', 'DLTKR', 'RTIOL', 'ERAIN']

# === Load Parameters from Master Excel ===
xls_master = pd.ExcelFile(param_master_file)
tc_df = xls_master.parse("Tc")
param_data = {param: xls_master.parse(param) for param in parameters}

# === Function to process each Excel file ===
def process_true_simulation(filename):
    if not filename.endswith(".xlsx"):
        return

    excel_path = os.path.join(excel_dir, filename)
    nam = os.path.splitext(filename)[0]
    name = nam[:2] + nam[3:6]

    try:
        point_id = int(name)
    except ValueError:
        print(f"[SKIPPED] Invalid Point_ID in {filename}")
        return

    print(f"\n🔄 Running TRUE simulation for: {name} (Point_ID: {point_id})")

    # Get TRUE Tc value
    try:
        tc_true = tc_df.loc[tc_df['Point_ID'] == point_id, 'True'].values[0]
    except IndexError:
        print(f"[WARNING] Tc TRUE value not found for Point_ID {point_id}")
        return

    # Update Geometry sheet with Tc TRUE value
    wb = load_workbook(excel_path)
    if 'Geometry' not in wb.sheetnames:
        print(f"[SKIPPED] Geometry sheet not found in {filename}")
        return
    ws = wb['Geometry']
    ws.cell(row=3, column=1).value = tc_true
    wb.save(excel_path)

    # Extract TRUE values for other parameters
    param_values = {}
    for param in parameters:
        df = param_data[param]
        row = df[df['Point_ID'] == point_id]
        if row.empty:
            print(f"[WARNING] TRUE value for {param} not found for Point_ID {point_id}")
            return
        param_values[param] = row['True'].values[0]

    # Folder name and path
    folder_name = "__".join([f"{p}_True" for p in parameters])
    subfolder_path = os.path.join(output_base_dir, name, folder_name)
    os.makedirs(subfolder_path, exist_ok=True)

    # Copy contents of exefiles folder
    for item in os.listdir(exefiles_dir):
        s = os.path.join(exefiles_dir, item)
        d = os.path.join(subfolder_path, item)
        if os.path.isdir(s):
            shutil.copytree(s, d, dirs_exist_ok=True)
        else:
            shutil.copy2(s, d)

    # Copy and rename Excel file
    shutil.copy2(excel_path, os.path.join(subfolder_path, "input.xlsx"))

    # Write input.bsn using TRUE values
    bsn_content = f"""1\t\t\t!hydrograph timestep(hr)
153.00\t\t!area (km2)
6.97
1.20
0.04\t\t!Snow melt coeff
0.02\t\t!Freezing temp (C)
{param_values['STRKR']:.2f}\t\t!LE1
{param_values['DLTKR']:.2f}\t\t!LE2
{param_values['RTIOL']:.2f}\t\t!LE3
{param_values['ERAIN']:.2f}\t\t!LE4
1.93\t\t!UC1 (TC)
{param_values['Tr']:.2f}\t\t!UC2
"""
    with open(os.path.join(subfolder_path, "input.bsn"), "w") as f:
        f.write(bsn_content)

    # Run the executable
    exe_path = os.path.join(subfolder_path, exe_name)
    try:
        subprocess.run([exe_path], cwd=subfolder_path, check=True)
        print(f"✅ TRUE simulation completed: {folder_name}")
    except Exception as e:
        print(f"[ERROR] Failed to run exe in {folder_name}: {e}")

# === Run in parallel ===
excel_files = [f for f in os.listdir(excel_dir) if f.endswith(".xlsx")]
Parallel(n_jobs=max(1, multiprocessing.cpu_count() - 1))(
    delayed(process_true_simulation)(filename)
    for filename in tqdm(excel_files, desc="Processing TRUE simulations", ncols=80)
)

print("\n✅ All TRUE simulations completed.")