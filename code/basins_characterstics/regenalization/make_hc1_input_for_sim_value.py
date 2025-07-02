#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jul 1 12:00:16 2025
@author: pouria
"""

import os
import shutil
import itertools
import subprocess
import pandas as pd
from openpyxl import load_workbook
from joblib import Parallel, delayed
import multiprocessing
from tqdm import tqdm

# === CONFIGURATION ===
excel_dir = "C:/Users/DrBZahraei/Desktop/regenalization/hydrographs"
exefiles_dir = "C:/Users/DrBZahraei/Desktop/regenalization/Edit-Hec-For_Dr_Zahraii2"
output_base_dir = "C:/Users/DrBZahraei/Desktop/regenalization/output"
param_master_file = "C:/Users/DrBZahraei/Desktop/regenalization/parameters.xlsx"
exe_name = "Read_input_xlsx_+_observ.exe"
parameters = ['Tr', 'STRKR', 'DLTKR', 'RTIOL', 'ERAIN']
sim_labels = ['Sim1', 'Sim2', 'Sim3']

# === Load Parameters from Master Excel ===
xls_master = pd.ExcelFile(param_master_file)
tc_df = xls_master.parse("Tc")
param_data = {param: xls_master.parse(param) for param in parameters}

# === Generate 3^5 combinations ===
combinations = list(itertools.product(sim_labels, repeat=5))

# === Function to process each simulation subfolder ===
def process_combination(combo, param_values, base_folder, excel_path):
    combo_dict = dict(zip(parameters, combo))
    folder_name = "__".join([f"{p}_{s}" for p, s in combo_dict.items()])
    subfolder_path = os.path.join(base_folder, folder_name)
    os.makedirs(subfolder_path, exist_ok=True)

    # Copy contents of exefiles folder
    for item in os.listdir(exefiles_dir):
        s = os.path.join(exefiles_dir, item)
        d = os.path.join(subfolder_path, item)
        if os.path.isdir(s):
            shutil.copytree(s, d, dirs_exist_ok=True)
        else:
            shutil.copy2(s, d)

    # Copy and rename Excel file
    shutil.copy2(excel_path, os.path.join(subfolder_path, "input.xlsx"))

    # Extract parameter values
    le1 = param_values['STRKR'][combo_dict['STRKR']]
    le2 = param_values['DLTKR'][combo_dict['DLTKR']]
    le3 = param_values['RTIOL'][combo_dict['RTIOL']]
    le4 = param_values['ERAIN'][combo_dict['ERAIN']]
    uc2 = param_values['Tr'][combo_dict['Tr']]

    # Write input.bsn
    bsn_content = f"""1\t\t\t!hydrograph timestep(hr)
                    153.00\t\t!area (km2)
                    6.97
                    1.20
                    0.04\t\t!Snow melt coeff
                    0.02\t\t!Freezing temp (C)
                    {le1:.2f}\t\t!LE1
                    {le2:.2f}\t\t!LE2
                    {le3:.2f}\t\t!LE3
                    {le4:.2f}\t\t!LE4
                    1.93\t\t!UC1 (TC)
                    {uc2:.2f}\t\t!UC2
                    """
    with open(os.path.join(subfolder_path, "input.bsn"), "w") as f:
        f.write(bsn_content)

    # Run the executable
    exe_path = os.path.join(subfolder_path, exe_name)
    try:
        subprocess.run([exe_path], cwd=subfolder_path, check=True)
    except Exception as e:
        print(f"[ERROR] Failed in {subfolder_path}: {e}")

# === Process Each Excel File ===
for filename in os.listdir(excel_dir):
    if not filename.endswith(".xlsx"):
        continue

    excel_path = os.path.join(excel_dir, filename)
    nam = os.path.splitext(filename)[0]
    name = nam[:2] + nam[3:6]
    print(f"\n🔄 Processing Excel file: {name}")

    # Identify Point_ID
    try:
        point_id = int(name)
    except ValueError:
        print(f"[SKIPPED] Could not parse Point_ID from {filename}")
        continue

    try:
        sim_value_tc = tc_df.loc[tc_df['Point_ID'] == point_id, 'Sim'].values[0]
    except IndexError:
        print(f"[WARNING] Point_ID {point_id} not found in Tc sheet of parameters.xlsx")
        continue

    # Update Geometry sheet in Excel file
    wb = load_workbook(excel_path)
    if 'Geometry' not in wb.sheetnames:
        print(f"[SKIPPED] Geometry sheet not found in {filename}")
        continue
    ws = wb['Geometry']
    ws.cell(row=3, column=1).value = sim_value_tc
    wb.save(excel_path)

    # Create base output folder
    base_folder = os.path.join(output_base_dir, name)
    os.makedirs(base_folder, exist_ok=True)

    # Extract parameter values for this Point_ID
    param_values = {}
    skip_flag = False
    for param in parameters:
        df = param_data[param]
        row = df[df['Point_ID'] == point_id]
        if row.empty:
            print(f"[WARNING] Point_ID {point_id} not found in {param} sheet.")
            skip_flag = True
            break
        param_values[param] = {
            'Sim1': row['Sim1'].values[0],
            'Sim2': row['Sim2'].values[0],
            'Sim3': row['Sim3'].values[0]
        }
    if skip_flag:
        continue

    # Run simulations in parallel with progress bar
    print(f"⚙️ Generating and executing 243 simulations for: {name}")
    Parallel(n_jobs=max(1, multiprocessing.cpu_count() - 1))(
        delayed(process_combination)(combo, param_values, base_folder, excel_path)
        for combo in tqdm(combinations, desc=f"Simulating {name}", ncols=80)
    )

print("\n✅ All files processed.")