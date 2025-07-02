#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel File Merger with MultiIndex Structure
Merges multiple Excel files using Point_ID as key and creates MultiIndex columns.

@author: adapted from your soilgrids processing code
"""

import os
import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook

# Set your input directory
input_path = r"/home/pouria/git/water-institute/data/basins_charactristics/output/excel/41/merege"
output_path = os.path.join(input_path, "merged_output.xlsx")

# Dictionary to hold data: {point_id: {filename: {column: value}}}
data = defaultdict(dict)

# Get all Excel files
excel_files = [f for f in os.listdir(input_path) if f.endswith(('.xlsx', '.xls'))]
print(f"Found {len(excel_files)} Excel files")

# Read all files and collect data
all_filenames = []
all_columns_per_file = {}

for file in excel_files:
    filename = os.path.splitext(file)[0]
    all_filenames.append(filename)
    file_path = os.path.join(input_path, file)
    
    try:
        df = pd.read_excel(file_path)
        
        # Standardize Point_ID column name (handle case sensitivity)
        if 'Point_Id' in df.columns:
            df = df.rename(columns={'Point_Id': 'Point_ID'})
            print(f"Renamed 'Point_Id' to 'Point_ID' in {filename}")
        
        if 'Point_ID' not in df.columns:
            print(f"Warning: No Point_ID column found in {filename}")
            continue
        
        # Get all columns except Point_ID
        data_columns = [col for col in df.columns if col != 'Point_ID']
        all_columns_per_file[filename] = data_columns
        
        print(f"Processing {filename}: {df.shape}, columns: {data_columns}")
        
        # Fill data dictionary
        for _, row in df.iterrows():
            point_id = row['Point_ID']
            if pd.isna(point_id):
                continue
                
            # Initialize point_id entry if it doesn't exist
            if filename not in data[point_id]:
                data[point_id][filename] = {}
            
            # Store all column values for this point_id and filename
            for col in data_columns:
                data[point_id][filename][col] = row[col]
                
    except Exception as e:
        print(f"Error processing {file}: {e}")

print(f"\nCollected data for {len(data)} unique Point_IDs")

# Get all unique point_ids and sort them
all_point_ids = sorted(data.keys())
all_filenames = sorted(all_filenames)

# Construct MultiIndex columns - similar to your soilgrids approach
columns = []
for filename in all_filenames:
    if filename in all_columns_per_file:
        for col in all_columns_per_file[filename]:
            columns.append((filename, col))

multi_columns = pd.MultiIndex.from_tuples(columns)

# Create DataFrame with MultiIndex columns
df_merged = pd.DataFrame(index=all_point_ids, columns=multi_columns)

# Fill DataFrame - similar to your approach
for point_id, file_data in data.items():
    for filename, col_data in file_data.items():
        for col_name, value in col_data.items():
            if (filename, col_name) in df_merged.columns:
                df_merged.loc[point_id, (filename, col_name)] = value

# Reset index to make Point_ID a column
df_merged = df_merged.rename(columns={'index': 'Point_ID'})

# Save to Excel with merged cells (like your soilgrids code)
df_merged.to_excel(output_path, merge_cells=True)

print(f"\nSuccess! Merged file saved to: {output_path}")
print(f"Final dataframe shape: {df_merged.shape}")
print(f"Number of Point_ID values: {len(df_merged)}")

# Display sample of the structure
print(f"\nSample of MultiIndex structure:")
print("Point_ID column:", df_merged.columns[0])
print("MultiIndex columns (first 10):")
for i, col in enumerate(df_merged.columns[1:]):
    if i < 10:
        print(f"  {col}")
    elif i == 10:
        print("  ...")
        break


# Load the existing Excel file
wb = load_workbook(output_path)

# Select the active worksheet (or specify by name)
ws = wb.active  # or wb['SheetName']


# Add two names in the first and second row, first column (A1 and A2)
ws['A1'] = 'Point_ID'  # Replace 'Name1' with your first name
ws['A2'] = 'Point_ID'  # Replace 'Name2' with your second name

# Delete the entire third row
ws.delete_rows(3)  # Deletes row number 3

# Save the changes
wb.save(output_path)