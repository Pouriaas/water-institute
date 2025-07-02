#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Merge Multiple MultiIndex Excel Files - Fixed Version
Properly handles Point_ID column when merging MultiIndex DataFrames.
"""

import os
import pandas as pd
from openpyxl import load_workbook

# Set your input directory and files
input_path = "/home/pouria/git/water-institute/data/basins_charactristics/output/excel/41/merege"

# List your MultiIndex Excel files here
multiindex_files = [
    "merged_output.xlsx",  # Your first merged file
    "soilgrids.xlsx",      # Your second file with MultiIndex
    # Add more files as needed
]

output_path = os.path.join(input_path, "final_merged_multiindex.xlsx")

print(f"Processing {len(multiindex_files)} MultiIndex files...")

dataframes = []

for file in multiindex_files:
    file_path = os.path.join(input_path, file)
    
    if not os.path.exists(file_path):
        print(f"Warning: {file} not found, skipping...")
        continue
    
    try:
        # Read Excel file with MultiIndex columns
        df = pd.read_excel(file_path, header=[0, 1])
        
        print(f"Processing {file}: {df.shape}")
        print(f"Columns: {list(df.columns)}")
        
        # Handle Point_ID column properly
        # Check if Point_ID exists as a non-MultiIndex column
        point_id_col = None
        
        # Look for Point_ID in different ways
        for col in df.columns:
            if col == 'Point_ID':  # Direct match
                point_id_col = col
                break
            elif isinstance(col, tuple) and col[0] == 'Point_ID':  # In MultiIndex
                point_id_col = col
                break
            elif isinstance(col, tuple) and 'Point_ID' in str(col):  # Contains Point_ID
                point_id_col = col
                break
        
        if point_id_col is not None:
            # Extract Point_ID as a separate Series
            point_ids = df[point_id_col].copy()
            
            # Remove Point_ID from the DataFrame
            df = df.drop(columns=[point_id_col])
            
            # Add Point_ID back as a regular column (not MultiIndex)
            df.insert(0, 'Point_ID', point_ids)
        else:
            print(f"Warning: No Point_ID column found in {file}")
            continue
        
        # Remove any rows where Point_ID is NaN or equals 'Point_ID' (header duplicates)
        df = df[df['Point_ID'].notna()]
        df = df[df['Point_ID'] != 'Point_ID']
        
        print(f"After cleaning: {df.shape}")
        print(f"Sample Point_IDs: {df['Point_ID'].head().tolist()}")
        
        dataframes.append(df)
        
    except Exception as e:
        print(f"Error processing {file}: {e}")
        import traceback
        traceback.print_exc()

if len(dataframes) == 0:
    print("No valid dataframes to merge!")
    exit()

# Start with the first dataframe
result = dataframes[0].copy()
print(f"\nStarting with DataFrame shape: {result.shape}")

# Merge with each subsequent dataframe
for i, df in enumerate(dataframes[1:], 1):
    print(f"Merging with DataFrame {i+1}, shape: {df.shape}")
    
    # Perform outer merge on Point_ID
    result = pd.merge(
        result, df, 
        on='Point_ID', 
        how='outer', 
        suffixes=('', f'_dup{i}')
    )
    
    print(f"After merge {i}: {result.shape}")

# Clean up any remaining issues
result = result[result['Point_ID'].notna()]
result = result[result['Point_ID'] != 'Point_ID']

# Sort by Point_ID for consistency
try:
    # Try to sort numerically if possible
    result['Point_ID_numeric'] = pd.to_numeric(result['Point_ID'], errors='coerce')
    if not result['Point_ID_numeric'].isna().all():
        result = result.sort_values('Point_ID_numeric').drop('Point_ID_numeric', axis=1)
    else:
        result = result.sort_values('Point_ID')
except:
    result = result.sort_values('Point_ID')

# Reset index
result = result.reset_index(drop=True)

# Save to Excel with merged cells
result.to_excel(output_path,  merge_cells=True)

print(f"\nSuccess! Final merged file saved to: {output_path}")
print(f"Final dataframe shape: {result.shape}")
print(f"Number of Point_ID values: {len(result)}")

# Display sample of the structure
print(f"\nFirst few rows:")
print(result.head())

print(f"\nColumn structure:")
print("Regular columns:", [col for col in result.columns if not isinstance(col, tuple)])
print("MultiIndex columns (first 10):", [col for col in result.columns if isinstance(col, tuple)][:10])

# Load the existing Excel file
wb = load_workbook(output_path)

# Select the active worksheet (or specify by name)
ws = wb.active  # or wb['SheetName']


# Add two names in the first and second row, first column (A1 and A2)
ws['A1'] = 'Point_ID'  # Replace 'Name1' with your first name
ws['A2'] = 'Point_ID'  # Replace 'Name2' with your second name

# Delete the entire third row
ws.delete_rows(3)  # Deletes row number 3

# Save the changes
wb.save(output_path)