#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jun  2 14:28:04 2025
Updated to fix extra row and drop empty data rows.

@author: pouria
"""

import os
import numpy as np
import pandas as pd
import rasterio
from collections import defaultdict
from openpyxl import load_workbook

# Path to the main directory containing feature folders
main_dir = "/home/pouria/git/water-institute/data/basins_charactristics/output/soilgrids_cliped/14"
output_path = "/home/pouria/git/water-institute/data/basins_charactristics/output/excel/14/soilgrids.xlsx"

# Dictionary to hold stats: {point_id: {feature: [mean, p5, p50, p95]}}
data = defaultdict(dict)

# Get list of feature folders
features = [f for f in os.listdir(main_dir) if os.path.isdir(os.path.join(main_dir, f))]

for feature in features:
    feature_path = os.path.join(main_dir, feature)
    for tif_file in os.listdir(feature_path):
        if tif_file.endswith(".tif"):
            point_id = os.path.splitext(tif_file)[0]
            tif_path = os.path.join(feature_path, tif_file)

            try:
                with rasterio.open(tif_path) as src:
                    array = src.read(1).astype(float)

                    # Remove -32768 and NaN values
                    array = array[(array != -32768) & ~np.isnan(array)]

                    # Scale valid values
                    array = array / 10.0

                    if array.size > 0:
                        mean = np.mean(array)
                        p5 = np.percentile(array, 5)
                        p50 = np.percentile(array, 50)
                        p95 = np.percentile(array, 95)
                    else:
                        mean = p5 = p50 = p95 = np.nan

                data[point_id][feature] = [mean, p5, p50, p95]
            except Exception as e:
                print(f"Error processing {tif_path}: {e}")

# Construct MultiIndex columns
all_point_ids = sorted(data.keys())
all_features = sorted(features)

columns = [(feature, stat) for feature in all_features for stat in ['mean', 'p5', 'p50', 'p95']]
multi_columns = pd.MultiIndex.from_tuples(columns)

df = pd.DataFrame(columns=multi_columns)

# Fill DataFrame
for point_id, stats_dict in data.items():
    for feature, values in stats_dict.items():
        for i, stat in enumerate(['mean', 'p5', 'p50', 'p95']):
            df.loc[point_id, (feature, stat)] = values[i]


# Save to Excel
df.to_excel(output_path, merge_cells=True)

# Load the existing Excel file
wb = load_workbook(output_path)

# Select the active worksheet (or specify by name)
ws = wb.active  # or wb['SheetName']

# Add two names in the first and second row, first column (A1 and A2)
ws['A1'] = 'Point_ID'  # Replace 'Name1' with your first name
ws['A2'] = 'Point_ID'  # Replace 'Name2' with your second name

# Delete the entire third row
ws.delete_rows(3)  # Deletes row number 3

# Save the changes
wb.save(output_path)
