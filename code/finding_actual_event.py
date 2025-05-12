import numpy as np
import pandas as pd
import os
import openpyxl

input_file = r'/home/pouria/git/water-institute/data/checking_similarity'

# Output Excel file path
output_excel_path = os.path.join(input_file, "output.xlsx")

# Load calendar
shamsi_calen = pd.read_excel(r"/home/pouria/git/water-institute/data/solar_hijri_dates_1350_to_1401.xlsx")

# Open an ExcelWriter to write multiple sheets
with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:

    files_and_directory = os.listdir(input_file)
    for folder in files_and_directory:
        path_folder = os.path.join(input_file, folder)
        if not os.path.isdir(path_folder):
            continue  # skip files
        
        stations_name = os.listdir(path_folder)
        
        res = [s.replace("_hourly.xlsx","") for s in stations_name]  
        res = [s.replace("-","") for s in res]  
        

        array = np.full((len(stations_name) + 1, len(shamsi_calen)), np.nan, dtype="object")
        df = pd.DataFrame(array.T, columns=["Date"] + res)
        df["Date"] = shamsi_calen
        df = df.set_index("Date")

        for file_name in stations_name:
            path = os.path.join(path_folder, file_name)
            if not file_name.endswith(".xlsx"):
                continue
            reame=file_name.replace("_hourly.xlsx","")
            reame=reame.replace("-","")
            wb = openpyxl.load_workbook(path)
            # in this time all the excel are green so we don't need remove by color
            # target_color = 'FF00B050'
            # purple_sheets = []

            # for sheetname in wb.sheetnames:
            #     sheet = wb[sheetname]
            #     tab_color = sheet.sheet_properties.tabColor

            #     if tab_color is not None:
            #         tab_color_rgb = tab_color.rgb
            #         if tab_color_rgb and tab_color_rgb.upper() == target_color:
            #             purple_sheets.append(sheetname)

            for sheet_name in wb.sheetnames:
                sim = pd.read_excel(path, sheet_name=sheet_name)
                valid_date = sim["Date"].dropna().unique()
                sim = sim.set_index("Date")
                for date in valid_date:
                    try:
                        if pd.isna(sim["Time"][date]):
                            continue
                        year = date.split("/")[0]
                        if year == "00":
                            year = "1400"
                        else:
                            year = "13" + str(year)
                        newformat = date.split("/")
                        new_date = f"{year}/{newformat[1]}/{newformat[2]}"
                        df.loc[new_date, reame] = 1
                    except:
                        continue  # Skip if indexing error

        # Add an 'Event' column to hold labels (like "--- GAP ---")
        df.insert(0, 'Event', '')
        
        # Define which rows qualify as part of an "event" (≥2 non-NaN values excluding 'Event')
        is_event_row = df.drop(columns='Event').notna().sum(axis=1) >= 2
        
        # Mark consecutive blocks of event rows
        group_id = (is_event_row != is_event_row.shift()).cumsum()
        
        # Group only the rows that are part of event blocks
        grouped_events = df[is_event_row].groupby(group_id)
        
        # Build final DataFrame with labeled gaps
        final_rows = []
        for _, group in grouped_events:
            final_rows.append(group)
            # Insert a labeled gap row
            gap_row = pd.DataFrame([['--- GAP ---'] + [np.nan] * (df.shape[1] - 1)],
                                   columns=df.columns,
                                   index=[pd.NaT])
            final_rows.append(gap_row)
        
        # Concatenate everything
        df = pd.concat(final_rows)
        
        # Optional: drop the very last "--- GAP ---" if it's the final row
        if df['Event'].iloc[-1] == '--- GAP ---':
            df = df.iloc[:-1]
        
        # Save to a new sheet with the name of the folder
        df.to_excel(writer, sheet_name=folder[:31])  # Sheet names must be ≤ 31 chars